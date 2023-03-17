import pandas as pd 
import openpyxl
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment 
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

list_rows_total = []

# Leer el archivo CSV
df = pd.read_csv('ReporteOAGUS_20230305.csv')

# Crear un nuevo libro de trabajo de Excel
wb = Workbook()

# Seleccionar la hoja activa
ws = wb.active

# Establecer el ancho de las columnas
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12 
ws.column_dimensions['C'].width = 5     #COLUMNA 1
ws.column_dimensions['D'].width = 8     #COLUMNA 2
ws.column_dimensions['E'].width = 5     #COLUMNA 3
ws.column_dimensions['G'].width = 10    #COLUMNA 4
ws.column_dimensions['H'].width = 5    #COLUMNA 5
ws.column_dimensions['I'].width = 5    #COLUMNA 6
ws.column_dimensions['J'].width = 5     #COLUMNA 7
ws.column_dimensions['K'].width = 5     #COLUMNA 8
ws.column_dimensions['L'].width = 5     #COLUMNA 9
ws.column_dimensions['M'].width = 5     #COLUMNA 10
ws.column_dimensions['N'].width = 6     #COLUMNA 11
ws.column_dimensions['O'].width = 8    #COLUMNA 12
ws.column_dimensions['P'].width = 10    #COLUMNA 13

# Establecer la altura deseada en la fila y columna especificada
ws.row_dimensions[1].height = 20


# Establecer el estilo de fuente y alineación
header_font = Font(name='Calibri', size=6, bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
cell_alignment = Alignment(horizontal='left', vertical='center')


# Crear una lista con los nombres de las cabeceras
cabeceras = ['CA', 'Market', 'Ind AM', 'Region', 'Month', 'Chg', 'Prev Ops', 'New Ops', 'Ops Chg', 'Prev Seat', 'New Seat', 'Sea Chg', '%_Seat Chg']

headers = cabeceras                             #Indico el número de columna a iniciar las cabeceras.
for col_num, header_title in enumerate(headers, 4):
    cell = ws.cell(row=1, column=col_num, value=header_title)
    cell.font = header_font
    cell.alignment = header_alignment



# Justifica el texto de los encabezados
for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    

# cambiar el formato de color de los renglones 
def apply_color(idx):
    # definir los colores para los renglones
    color = 'F2F2F2' if idx % 2 == 0 else 'FFFFFF'
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def apply_font_bold(value):
    return Font(name='Calibri', size=6, bold= True) if value == 'TOTAL' else Font(name='Calibri', size=6)

ca_actual = ''

# Escribir los datos                
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 3):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = apply_font_bold(row_data[2])
        cell.alignment = cell_alignment
        cell.fill = apply_color(row_num)


# Cambiar el color de la celda A1 a rojo

#fill = PatternFill(start_color='808080', end_color='FFFFFF', fill_type='solid')

# Iterar sobre todas las filas y aplicar el formato de relleno al patrón especificado
#for row in ws.iter_rows():
 #   if row[0].value == 'TOTAL':
  #      cell_font = Font(bold = True)
   #     for cell in row:
    #        cell.fill = fill

#color1 = 'BFBFBF' COLOR DE LA CABECERA


# Seleccionar las celdas que se van a ajustar
#cell_range = ws['I1:M1']

# Ajustar el ancho de las columnas para que el texto quepa
#for row in cell_range:
 #   for cell in row:
  #      ws.column_dimensions[cell.column_letter].width = len(str(cell.value))

# Justifica el texto de los encabezados
for cell in ws[1]:
    cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')    #Agrego el color de relleno en el renglon 1


#Convirtiendo a Decimales
# Selecciona la columna que deseas convertir (por ejemplo, columna A)
columna = ws['P']

# Itera sobre las celdas en la columna y convierte los valores de decimal a porcentaje
for celda in columna:
    if isinstance(celda.value, float):  # verifica si el valor de la celda es un decimal
        celda.value = celda.value * 1  # convierte el valor a porcentaje
        celda.number_format = '0%'  # establece el formato de número de la celda como porcentaje con dos decimales



# Crea un objeto estilo para aplicar a la columna 'N'
#font = Font(color='FF0000')
#fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Itera sobre cada fila de la columna 'M'
#for index, row in df.iterrows():
 #   if row['M'] < 0:  # Si el valor en la columna 'M' es negativo
  #      # Aplica el formato deseado a la celda en la columna 'N'
 #  3     cell = f'{get_column_letter(df.columns.get_loc("N") + 1)}{index + 1}'
#    df.loc[index, 'N'] = f'({abs(row["N"])}%)'
  #      df[cell].font = font
 #       df[cell].fill = fill

# Guarda el archivo xlsx con los cambios realizados
#df.to_excel('tu_archivo_modificado.xlsx', index=False)

# Indicar el número de columna que deseas eliminar
num_columna = 3
num_columna17 = 17
num_columna18 = 16
num_columna19 = 12

# Eliminar la columnas
ws.delete_cols(num_columna)
ws.delete_cols(num_columna17)
ws.delete_cols(num_columna18)
ws.delete_cols(num_columna19)

# Guardar el archivo
wb.save('archivo.xlsx')


